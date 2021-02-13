"""
Excel et python

Dans ce programme on charge un classeur Excel, on apprend à ajouter une liste de données dans le fichier concerné

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Assignation d'une liste de noms"""
names = ['Dan', 'April', 'Nea']

"""Ajout de données"""
starting_row = 12  # assignation de la ligne de départ d'ajout de données
for name in names:
    ws.cell(row=starting_row, column=1).value = name  # ajout à la colonne 1 à partir de la ligne 12
    starting_row += 1  # incrémentation +1 des lignes à ajouter les données

"""Sauvegarde des données du classeur concerné"""
wb.save('pieces/name_color.xlsx')
print('Le fichier a été sauvegardé')
