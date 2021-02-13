"""
Excel et python

Dans ce programme on apprend intervenir sur la police d'écriture

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import Workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant
from openpyxl.styles import Font  # changement de la police d'écriture

"""Chargement du classeur Excel (classeur = workbook = wb)"""
wb = load_workbook('pieces/names.xlsx')

"""Création d'une feuille"""
ws = wb.active

"""Assignation de la sélection de cellules"""
cell1 = ws['A1']
cell2 = ws['B1']
cell3 = ws['C1']

"""Changement de la police d'écriture"""
cell1.font = Font(
    size=15,
    bold=True,
    italic=False,
    color='253bb8'
)
cell2.font = Font(
    size=20,
    bold=False,
    italic=True,
    color='253bb8'
)
cell3.font = Font(
    size=20,
    bold=True,
    italic=True,
    color='253bb8'
)

"""Sauvegarde de la feuille"""
wb.save('pieces/names.xlsx')
print('Fichier sauvegardé !')
