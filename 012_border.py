"""
Excel et python

Dans ce programme on apprend à intervenir sur la configuration de la bordure des cellules

Liste des différents types de bordure :
'thin', 'dashed', 'dashDotDot', 'hair', 'thick', 'dotted', 'mediumDashDotDot', 'dashDot', 'double', 'mediumDashDot',
'mediumDashed', 'slantDashDot', 'medium'

Éditeur : Laurent REYNAUD
Date : 01-02-2021
"""

from openpyxl.workbook import Workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant
from openpyxl.styles import Font, Border, Side  # changement de la police d'écriture

"""Chargement du classeur Excel (classeur = workbook = wb)"""
wb = load_workbook('pieces/names.xlsx')

"""Création d'une feuille"""
ws = wb.active

"""Assignation de la sélection de cellules"""
cell1 = ws['A1']
cell2 = ws['B1']
cell3 = ws['C1']

"""Changement de la police d'écriture"""
cell1.font = Font(
    size=15,
    bold=True,
    italic=False,
    color='253bb8'
)
cell2.font = Font(
    size=20,
    bold=False,
    italic=True,
    color='253bb8'
)
cell3.font = Font(
    size=20,
    bold=True,
    italic=True,
    color='253bb8'
)

"""Configuration d'un côté de la bordure"""
my_bd = Side(style='thick', color='d80d0d')  # thick = épais

"""Assignation de la cellule B3"""
B3 = ws['B3']

"""Bordure à appliquer à la cellule B3"""
B3.border = Border(left=my_bd,
                   right=my_bd,
                   top=my_bd,
                   bottom=my_bd)

"""Configuration d'un côté de la bordure"""
my_bd2 = Side(style='double', color='000000')  # thick = épais

"""Bordure à appliquer aux cellules de la ligne 1"""
cell1.border = Border(bottom=my_bd2)
cell2.border = Border(bottom=my_bd2)
cell3.border = Border(bottom=my_bd2)

"""Sauvegarde de la feuille"""
wb.save('pieces/names.xlsx')
print('Fichier sauvegardé !')
