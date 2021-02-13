"""
Excel et python

Dans ce programme on charge un classeur Excel, on apprend à ajouter les données dans le fichier concerné

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Ajout de données"""
starting_row = 11  # assignation de la ligne d'ajout de données
ws.cell(row=starting_row, column=1).value = 'Neo'  # ajout à la ligne 11 colonne 1
ws.cell(row=starting_row, column=2).value = 'Noir'  # ajout à la ligne 11 colonne 2

"""Sauvegarde des données du classeur concerné"""
wb.save('pieces/name_color.xlsx')
print('Le fichier a été sauvegardé')
