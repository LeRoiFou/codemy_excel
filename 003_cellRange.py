"""
Excel et python

Dans ce programme on charge un classeur Excel et on récupère une plage de cellules

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Saisie d'une plage de cellules"""
range = ws['A2':'B11']

"""Affichage des données"""
print(range)  # affichage d'un tuple
for cell in range:
    for x in cell:  # double boucle cette fois-ci
        print(x.value)
