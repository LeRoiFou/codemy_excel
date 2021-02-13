"""
Excel et python

Dans ce programme on charge un classeur Excel, on change les données d'une cellule et on enregistre le fichier Excel

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Changement des données d'une cellule"""
ws['A2'] = 'Laurent'

"""Sauvegarde des données du classeur concerné"""
wb.save('pieces/name_color.xlsx')
print('Le fichier a été sauvegardé')
