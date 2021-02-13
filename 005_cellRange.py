"""
Excel et python

Dans ce programme on charge un classeur Excel et on récupère une plage de cellules par colonne selon une deuxième
méthode : l'itération, avec la méthode prédéfinie iter_cols()
Dans l'exemple ci-après le résultat est le même qu'avec la méthode prédéfinie iter_rows().

On a donc deux méthodes pour récupérer les données d'une plage de cellule :
-> méthode 1 : saisie d'une plage de cellule (excel-3)
-> méthode 2 : recours aux méthodes préétablies iter_rows() / iters_cols()

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Récupération et affichage d'une plage de cellules"""
for col in ws.iter_cols(min_row=2, max_row=6, min_col=1, max_col=1, values_only=True):
    for cell in col:  # deuxième boucle pour extraire les données du tuple
        print(cell)
