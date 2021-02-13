"""
Excel et python

Dans ce programme on charge un classeur Excel et on récupère les données d'une colonne / d'une ligne

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Récupération des données de la colonne A"""
column_a = ws['A']

"""Affichage des données"""
print(column_a)  # affichage d'un tuple
for cell in column_a:
    print(cell.value)

"""Récupération des données de la ligne 1"""
line_1 = ws['1']

"""Affichage des données"""
for cell in line_1:
    print(cell.value)
