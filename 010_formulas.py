"""
Excel et python

Dans ce programme on apprend à utiliser les formules Excel dans Python

ATTENTION LES FORMULES SONT EN ANGLAIS !!!

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import Workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

"""Création d'un classeur Excel (classeur = workbook = wb)"""
wb = Workbook()

"""Création d'une feuille"""
ws = wb.active

"""Assignation d'une liste de noms, d'une liste de couleurs et d'une liste de nombre"""
names = ['Dan', 'April', 'Neal', 'Sara']
colors = ['Bleu', 'Violet', 'Vert', 'Blanc']
nums = [12, 39, 42, 21]

"""Ajout des titres à la 1ère ligne"""
ws['A1'] = 'Nom'
ws['B1'] = 'Couleur'
ws['C1'] = 'Nombre favori'

"""Ajout des noms à partir de la 2ème ligne dans la colonne n° 1"""
starting_row = 2  # ligne de départ
for name in names:
    ws.cell(row=starting_row, column=1).value = name
    starting_row += 1  # incrémentation de +1

"""Ajout des couleurs à partir de la 2ème ligne dans la colonne n° 2"""
starting_row = 2  # ligne de départ
for color in colors:
    ws.cell(row=starting_row, column=2).value = color
    starting_row += 1  # incrémentation de +1

"""Ajout des nombres à partir de la 2ème ligne dans la colonne n° 3"""
starting_row = 2  # ligne de départ
for number in nums:
    ws.cell(row=starting_row, column=3).value = number
    starting_row += 1  # incrémentation de +1

"""Ajout de formules"""
ws['C6'] = '=SUM(C2:C5)'  # ATTENTION LA FONCTION EST EN ANGLAIS !!!
ws['C7'] = '=AVERAGE(C2:C5)'  # ATTENTION LA FONCTION EST EN ANGLAIS !!!

"""Ajout d'un nom de l'onglet"""
ws.title = 'Noms_couleurs_chiffres'

"""Sauvegarde de la feuille"""
wb.save('pieces/names.xlsx')
print('Fichier sauvegardé !')
