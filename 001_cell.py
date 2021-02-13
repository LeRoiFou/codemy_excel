"""
Excel et python

Dans ce programme on charge un classeur Excel et on récupère une donnée dans une cellule

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Assignation d'un nom rattachée à la cellule Excel concernée"""
name = ws['A5'].value

"""Assignation d'une couleur rattachée à la cellue Excel concernée"""
color = ws['B5'].value

"""Affichage des données dans les cellules sélectionnées"""
print(f"{name} : {color}")
