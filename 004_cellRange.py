"""
Excel et python

Dans ce programme on charge un classeur Excel et on récupère une plage de cellules par ligne selon une deuxième
méthode : l'itération, avec la méthode prédéfinie iter_rows()

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

# """Création d'un classeur Excel (classeur = workbook = wb)"""
# wb = workbook()

"""Chargement de la feuille active d'Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation de la feuille active d'Excel (feuille = worksheet = ws)"""
ws = wb.active

"""Récupération et affichage d'une plage de cellules"""
for row in ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=1, values_only=True):
    for cell in row:  # deuxième boucle pour extraire les données du tuple
        print(cell)
