"""
Excel et python

Dans ce programme on créé un fichier Excel en ajoutant un nom à l'onglet, et à ajouter les données

Éditeur : Laurent REYNAUD
Date : 28-01-2021
"""

from openpyxl.workbook import Workbook  # création d'une classeur Excel
from openpyxl import load_workbook  # chargement d'un classeur Excel existant

"""Création d'un classeur Excel (classeur = workbook = wb)"""
wb = Workbook()

"""Création d'une feuille"""
ws = wb.active

"""Assignation d'une liste de noms et d'une liste de couleurs"""
names = ['Dan', 'April', 'Neal', 'Sara']
colors = ['Bleu', 'Violet', 'Vert', 'Blanc']

"""Ajout des titres à la 1ère ligne"""
ws['A1'] = 'Nom'
ws['B1'] = 'Couleur'

"""Ajout des noms à partir de la 2ème ligne dans la colonne n° 1"""
starting_row = 2  # ligne de départ
for name in names:
    ws.cell(row=starting_row, column=1).value = name
    starting_row += 1  # incrémentation de +1

"""Ajout des couleurs à partir de la 2ème ligne dans la colonne n° 2"""
starting_row = 2  # ligne de départ
for color in colors:
    ws.cell(row=starting_row, column=2).value = color
    starting_row += 1  # incrémentation de +1

"""Ajout d'un nom de l'onglet"""
ws.title = 'Noms et couleurs'

"""Sauvegarde de la feuille"""
wb.save('pieces/names.xlsx')
print('Fichier sauvegardé !')
